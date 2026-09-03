import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests import _env  # noqa: F401
from tests import _stub_gcp
_stub_gcp.install()

from delivery.excel_export import build_repayment_workbook, build_sick_leave_workbook
from delivery.repository import repayment_matches_filters, sick_leave_matches_filters


class RepaymentMatchesFiltersTests(unittest.TestCase):
    def _record(self, **overrides):
        base = {"personnel_name": "王小明", "vendor": "shopee", "occurred_date": "2026-03-15"}
        base.update(overrides)
        return base

    def test_no_filters_matches(self):
        self.assertTrue(repayment_matches_filters(self._record()))

    def test_name_keyword_excludes_non_matching(self):
        self.assertFalse(repayment_matches_filters(self._record(), name_keyword="李小華"))

    def test_name_keyword_matches_substring(self):
        self.assertTrue(repayment_matches_filters(self._record(), name_keyword="小明"))

    def test_vendor_filter_excludes_non_matching(self):
        self.assertFalse(repayment_matches_filters(self._record(), vendor_filter="ud"))

    def test_vendor_filter_matches(self):
        self.assertTrue(repayment_matches_filters(self._record(), vendor_filter="shopee"))

    def test_month_filter_matches(self):
        self.assertTrue(repayment_matches_filters(self._record(), month_filter="2026-03"))

    def test_month_filter_excludes_non_matching_month(self):
        self.assertFalse(repayment_matches_filters(self._record(), month_filter="2026-04"))


class SickLeaveMatchesFiltersTests(unittest.TestCase):
    def _record(self, **overrides):
        base = {
            "personnel_name": "王小明",
            "vendor": "shopee",
            "start_date": "2026-03-15",
            "end_date": "2026-03-16",
            "leave_type": "sick",
        }
        base.update(overrides)
        return base

    def test_no_filters_matches(self):
        self.assertTrue(sick_leave_matches_filters(self._record()))

    def test_name_keyword_excludes_non_matching(self):
        self.assertFalse(sick_leave_matches_filters(self._record(), name_keyword="李小華"))

    def test_vendor_filter_excludes_non_matching(self):
        self.assertFalse(sick_leave_matches_filters(self._record(), vendor_filter="ud"))

    def test_month_filter_matches_start_date(self):
        self.assertTrue(sick_leave_matches_filters(self._record(), month_filter="2026-03"))

    def test_month_filter_excludes_non_matching_month(self):
        self.assertFalse(sick_leave_matches_filters(self._record(), month_filter="2026-04"))

    def test_leave_type_filter_matches(self):
        self.assertTrue(sick_leave_matches_filters(self._record(), leave_type_filter="sick"))

    def test_leave_type_filter_excludes_non_matching(self):
        self.assertFalse(sick_leave_matches_filters(self._record(), leave_type_filter="annual"))


class ExcelExportTests(unittest.TestCase):
    def test_repayment_workbook_contains_expected_rows(self):
        records = [
            {
                "occurred_date": "2026-03-15",
                "vendor": "shopee",
                "personnel_name": "王小明",
                "amount": 500,
                "reason": "遺失商品",
                "approved": True,
            }
        ]
        content = build_repayment_workbook(records)
        self.assertTrue(content.startswith(b"PK"))  # .xlsx 是 zip 格式，開頭一定是 PK

        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]
        self.assertEqual(header, ["日期", "廠商", "人員", "金額", "原因", "核准狀態"])
        self.assertEqual(row, ["2026-03-15", "蝦皮", "王小明", 500, "遺失商品", "已核准"])

    def test_sick_leave_workbook_contains_expected_rows(self):
        records = [
            {
                "start_date": "2026-03-15",
                "end_date": "2026-03-16",
                "leave_type": "sick",
                "vendor": "ud",
                "personnel_name": "李小華",
                "reason": "感冒",
                "approved": False,
            }
        ]
        content = build_sick_leave_workbook(records)

        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]
        self.assertEqual(header, ["開始日期", "結束日期", "假別", "廠商", "人員", "原因", "核准狀態"])
        self.assertEqual(row, ["2026-03-15", "2026-03-16", "病假", "UD", "李小華", "感冒", "未核准"])


if __name__ == "__main__":
    unittest.main()
